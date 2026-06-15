import openpyxl
import pandas as pd
import json
import os
import re
from datetime import datetime

# ==========================================
# 1. CONFIGURATION
# ==========================================
MD_FILE = 'statement.md'
EXCEL_FILE = 'bookkeeping.xlsx'
SHEET_NAME = 'Credit Card'
KEYWORDS_FILE = 'keywords.json'
STATEMENT_YEAR = 2025

# ==========================================
# 2. KEYWORD BANK (loaded from JSON)
# ==========================================
def load_keywords():
    """Load keyword bank from JSON file. Create with defaults if missing."""
    if not os.path.exists(KEYWORDS_FILE):
        default = {
            "COSTCO": ["Costco", "Supplies"],
            "STAPLES": ["Staples", "Office expenses"],
            "SHELL": ["Shell", "Gas"],
            "ESSO": ["Esso", "Gas"],
            "PETRO": ["Petro-Canada", "Gas"],
            "UBER": ["Uber", "Travel"],
            "AIR CANADA": ["Air Canada", "Travel"],
            "MCDONALDS": ["McDonalds", "Meals & entertainment"],
            "STARBUCKS": ["Starbucks", "Meals & entertainment"],
            "TIM HORTONS": ["Tim Hortons", "Meals & entertainment"],
            "ROGERS": ["Rogers", "Cell phone"],
            "BELL": ["Bell", "Telephone & utilities"],
            "PAYMENT": ["Credit Card Payment", "Payment"],
            "THANK YOU": ["Credit Card Payment", "Payment"]
        }
        with open(KEYWORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(default, f, indent=2)
        print(f"Created {KEYWORDS_FILE} with default entries.")
        return default

    with open(KEYWORDS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_keywords(keyword_bank):
    """Save keyword bank back to JSON file."""
    with open(KEYWORDS_FILE, 'w', encoding='utf-8') as f:
        json.dump(keyword_bank, f, indent=2)

# ==========================================
# 3. EXCEL COLUMN MAP
# ==========================================
COLS = {
    'Date': 1, 'Description': 2, 'Expenses': 3, 'Pay_Back': 4, 'Ask_Accountant': 6,
    'Categories': {
        'Furniture': 7, 'Bank Charges': 8, 'Advertising': 9, 'Training': 10,
        'Business tax': 11, 'Delivery & Freight': 12, 'Gas': 13, 'Parking': 14,
        'Insurance': 15, 'Interest': 16, 'Maintenance & repair': 17,
        'Reimbursement': 18, 'Meals & entertainment': 19, 'Office expenses': 20,
        'Personal': 21, 'Supplies': 22, 'Legal & accounting': 23, 'Rent': 24,
        'Salaries': 25, 'Travel': 26, 'Telephone & utilities': 27, 'Cell phone': 28
    }
}

# ==========================================
# 4. HELPER FUNCTIONS
# ==========================================
MONTH_MAP = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
}

DATE_PATTERN = re.compile(
    r'(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+\d{1,2}',
    re.IGNORECASE
)

AMOUNT_PATTERN = re.compile(
    r'-?\$[\d,]+\.\d{2}|\([\d,]+\.\d{2}\)'
)

PREV_BALANCE_PATTERN = re.compile(
    r'PREVIOUS\s+STATEMENT\s+BALANCE.*?\$([\d,]+\.\d{2})',
    re.IGNORECASE
)

# Words to strip from descriptions (provinces, cities, noise)
STRIP_WORDS = [
    # Canadian provinces/territories
    'AB', 'BC', 'MB', 'NB', 'NL', 'NS', 'NT', 'NU', 'ON', 'PE', 'QC', 'SK', 'YT',
    # Common prefixes to strip
    'IN',
    # Common Canadian cities (Ontario/GTA focus)
    'AURORA', 'BRAMPTON', 'BRANDON', 'CALGARY', 'CHARLOTTETOWN', 'EDMONTON',
    'FREDERICTON', 'HALIFAX', 'HAMILTON', 'KITCHENER', 'LONDON', 'MARKHAM',
    'MISSISSAUGA', 'MONCTON', 'MONTREAL', 'NIAGARA', 'NORTH YORK', 'OSHAWA',
    'OTTAWA', 'QUEBEC', 'REGINA', 'RICHMOND', 'RICHMOND HILL', 'SASKATOON',
    'SCARBOROUGH', 'ST. CATHARINES', 'SUDBURY', 'THUNDER BAY', 'TORONTO',
    'VANCOUVER', 'VICTORIA', 'WINDSOR', 'WINNIPEG', 'YORK',
    # More GTA cities
    'UNIONVILLE', 'THORNHILL', 'VAUGHAN', 'PICKERING', 'AJAX', 'WHITBY',
    'OSHAWA', 'BOWMANVILLE', 'CLARINGTON', 'NEWMARKET', 'EAST GWILLIMBURY',
    'KING CITY', 'KING', 'SCHOMBERG', 'HOLLAND LANDING',
    'HILL', 'HILLON', 'NMARKHAM',  # Richmond Hill ON, N Markham
    # Common noise words
    'STORE', 'OUTLET', 'LOCATION',
]


def clean_description(desc):
    """Strip cities, provinces, numbers, and noise from raw description."""
    desc = str(desc).strip()

    # Remove phone numbers (various formats)
    desc = re.sub(r'\d{3}[-.]?\d{3}[-.]?\d{4}', '', desc)
    desc = re.sub(r'\(\d{3}\)\s*\d{3}[-.]?\d{4}', '', desc)

    # Remove store/transaction numbers (long digit strings)
    desc = re.sub(r'\b\d{5,}\b', '', desc)

    # Remove store numbers like #1234 or #12345
    desc = re.sub(r'#\d+', '', desc)

    # Remove asterisks and surrounding special chars
    desc = re.sub(r'\\?\*+', '', desc)

    # Remove forward slashes
    desc = re.sub(r'/', ' ', desc)

    # Remove backslashes
    desc = re.sub(r'\\', '', desc)

    # Remove remaining special chars (keep spaces, apostrophes, and alphanumeric)
    desc = re.sub(r"[^\w\s']", ' ', desc)

    # Clean up multiple spaces
    desc = re.sub(r'\s+', ' ', desc).strip()

    # Split into words and filter
    words = desc.split()
    cleaned_words = []
    for word in words:
        # Skip pure numbers
        if re.match(r'^\d+$', word):
            continue
        # Skip province codes (case-insensitive)
        if word.upper() in STRIP_WORDS:
            continue
        cleaned_words.append(word)

    result = ' '.join(cleaned_words).strip()

    # Title case for readability
    if result:
        result = result.title()

    return result if result else desc.title()

# Rows that look like transactions but are actually summary/info lines
SKIP_PATTERNS = [
    re.compile(r'SUBTOTAL', re.IGNORECASE),
    re.compile(r'NEW BALANCE', re.IGNORECASE),
    re.compile(r'PREVIOUS STATEMENT BALANCE', re.IGNORECASE),
    re.compile(r'PAYMENT DUE DATE', re.IGNORECASE),
    re.compile(r'CREDIT LIMIT', re.IGNORECASE),
    re.compile(r'AVAILABLE CREDIT', re.IGNORECASE),
    re.compile(r'MINIMUM PAYMENT', re.IGNORECASE),
    re.compile(r'Interest\s+Rate', re.IGNORECASE),
    re.compile(r'Fee[s]?\b', re.IGNORECASE),
    re.compile(r'Cash advances', re.IGNORECASE),
    re.compile(r'Purchases\s+&\s+debits', re.IGNORECASE),
    re.compile(r'Payments\s+&\s+credits', re.IGNORECASE),
    re.compile(r'Points earned', re.IGNORECASE),
    re.compile(r'New points balance', re.IGNORECASE),
    re.compile(r'Previous Points balance', re.IGNORECASE),
    re.compile(r'Redemption', re.IGNORECASE),
    re.compile(r'BALANCE\b', re.IGNORECASE),
    re.compile(r'P\.?O\.?\s+BOX', re.IGNORECASE),
]


def get_company_and_category(description, keyword_bank):
    """Clean description first, then match to get the Clean Name and the Category."""
    cleaned = clean_description(description)
    desc_upper = cleaned.upper()
    for keyword, (clean_name, category) in keyword_bank.items():
        if keyword in desc_upper:
            return clean_name, category
    return cleaned, 'Ask Accountant'


def parse_date(date_str, year):
    """Convert string dates like 'MAY 14' or '04/06/2025' into date objects."""
    date_str = str(date_str).strip()

    # Try MM/DD/YYYY format
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', date_str)
    if m:
        return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()

    # Try MON DD format (e.g., 'MAY 14')
    m = re.match(r'([A-Za-z]+)\s+(\d{1,2})', date_str)
    if m:
        month_str = m.group(1).upper()
        day = int(m.group(2))
        if month_str in MONTH_MAP:
            return datetime(year, MONTH_MAP[month_str], day).date()

    # Fallback: let pandas try
    if not re.search(r'\d{4}', date_str):
        date_str = f"{date_str} {year}"
    return pd.to_datetime(date_str).date()


def parse_amount(amt_str):
    """Convert amount strings like '$12.50', '-$50.00', '(25.00)' into floats."""
    if not amt_str:
        return 0.0
    amt_str = str(amt_str).replace('$', '').replace(',', '').strip()
    if amt_str.startswith('(') and amt_str.endswith(')'):
        amt_str = '-' + amt_str[1:-1]
    return float(amt_str)


def is_date_cell(text):
    """Check if a cell value can be parsed as a date."""
    text = str(text).strip()
    if not text:
        return False
    # Check for MON DD patterns
    if DATE_PATTERN.search(text):
        return True
    # Check for MM/DD/YYYY
    if re.match(r'\d{1,2}/\d{1,2}/\d{4}', text):
        return True
    return False


def is_amount_cell(text):
    """Check if a cell value can be parsed as a currency amount."""
    text = str(text).strip()
    if not text:
        return False
    return bool(AMOUNT_PATTERN.search(text))


def infer_columns_from_row(cells):
    """
    Given a list of cell values from the first data row, infer which
    cell index is the date, description, and amount using heuristics.
    Returns (date_idx, desc_idx, amount_idx) or None if inference fails.
    """
    date_idx = None
    amount_idx = None

    for i, cell in enumerate(cells):
        cell_text = str(cell).strip()
        if not cell_text:
            continue

        if date_idx is None and is_date_cell(cell_text):
            date_idx = i
        elif amount_idx is None and is_amount_cell(cell_text):
            amount_idx = i

    # Need at least a date and an amount
    if date_idx is None or amount_idx is None:
        return None

    # Description is whichever non-empty cell is not date or amount
    desc_idx = None
    for i, cell in enumerate(cells):
        if i == date_idx or i == amount_idx:
            continue
        cell_text = str(cell).strip()
        if cell_text:
            desc_idx = i
            break

    if desc_idx is None:
        # If no distinct description cell, description is embedded with date
        # Use the date cell as description source (will be cleaned later)
        desc_idx = date_idx

    return date_idx, desc_idx, amount_idx


def extract_transactions_from_line(line):
    """
    Extract transaction data from a single line of the markdown file.
    Returns (raw_date, raw_description, raw_amount) or None if not a transaction.
    """
    # Skip divider lines
    if re.match(r'^[\s|:-]+$', line):
        return None

    # Find all date matches
    dates = DATE_PATTERN.findall(line)
    if not dates:
        return None

    # Find all amount matches
    amounts = AMOUNT_PATTERN.findall(line)
    if not amounts:
        return None

    # Extract description: everything between the last date BEFORE the amount and the first amount
    # Find the last date that appears before the first amount
    first_amount_match = AMOUNT_PATTERN.search(line)
    last_date_match = None
    for m in DATE_PATTERN.finditer(line):
        if m.end() <= first_amount_match.start():
            last_date_match = m

    if last_date_match and first_amount_match:
        between = line[last_date_match.end():first_amount_match.start()].strip()
        # Clean up pipe characters and extra whitespace
        between = re.sub(r'\|+', ' ', between).strip()
        # Remove leading/trailing non-alphanumeric characters
        between = re.sub(r'^[\s|]+|[\s|-]+$', '', between).strip()
        raw_description = between if between else dates[0]
    else:
        raw_description = dates[0]

    # Skip lines that match known non-transaction patterns in the DESCRIPTION only
    # (supplementary info after the amount is OK - e.g., "Points earned")
    for pattern in SKIP_PATTERNS:
        if pattern.search(raw_description):
            return None

    # Skip lines where description is just a date (e.g., payment summary lines)
    if DATE_PATTERN.fullmatch(raw_description.strip()):
        return None

    return dates[0], raw_description, amounts[0]


def extract_all_transactions(md_file):
    """
    Read the markdown file and extract all transaction rows.
    Handles multi-page statements by scanning all lines.
    Returns list of (raw_date, raw_description, raw_amount) tuples.
    """
    with open(md_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    transactions = []
    for line in lines:
        line = line.strip()
        if not line or '|' not in line:
            continue

        result = extract_transactions_from_line(line)
        if result:
            transactions.append(result)

    return transactions


def extract_previous_balance(md_file):
    """Extract the previous statement balance from the markdown file."""
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # Look for "PREVIOUS STATEMENT BALANCE" followed by an amount
    match = PREV_BALANCE_PATTERN.search(content)
    if match:
        amt_str = match.group(1).replace(',', '')
        return float(amt_str)

    # Fallback: look for $750.85 or similar on a line with PREVIOUS
    for line in content.split('\n'):
        if 'PREVIOUS' in line.upper() and 'BALANCE' in line.upper():
            amounts = AMOUNT_PATTERN.findall(line)
            if amounts:
                return parse_amount(amounts[0])

    return None


# ==========================================
# 5. THE AUTOMATION ENGINE
# ==========================================
def parse_and_export():
    # --- STEP A: Load Keyword Bank ---
    keyword_bank = load_keywords()
    new_merchants = set()

    # --- STEP B: Extract Previous Balance ---
    prev_balance = extract_previous_balance(MD_FILE)
    if prev_balance is not None:
        print(f"Previous Balance: ${prev_balance:,.2f}")
    else:
        print("Warning: Could not extract Previous Balance from statement.")

    # --- STEP C: Extract Transactions from Markdown ---
    transactions = extract_all_transactions(MD_FILE)
    if not transactions:
        print("Error: No transactions found in the markdown file.")
        return

    print(f"Found {len(transactions)} candidate transactions.")

    # --- STEP D: Dynamic Column Inference ---
    first_raw = transactions[0]
    cells = [first_raw[0], first_raw[1], first_raw[2]]
    inferred = infer_columns_from_row(cells)

    if inferred:
        date_idx, desc_idx, amount_idx = inferred
        print(f"Inferred columns: date_idx={date_idx}, desc_idx={desc_idx}, amount_idx={amount_idx}")
    else:
        date_idx, desc_idx, amount_idx = 0, 1, 2
        print("Using default column order: date=0, desc=1, amount=2")

    # --- STEP E: Load Excel Template ---
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb[SHEET_NAME]
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # --- STEP F: Clear old data rows (Row 12+) ---
    for r in range(12, 200):
        for c in range(1, 30):
            sheet.cell(row=r, column=c).value = None
        # Stop when we hit a row that has a formula (row 10 is formulas, but we clear from 12)
        # We clear all data rows unconditionally

    # --- STEP G: Write Previous Balance Row (Row 12) ---
    DATA_START_ROW = 12
    if prev_balance is not None:
        sheet.cell(row=DATA_START_ROW, column=COLS['Description']).value = 'Previous Balance'
        sheet.cell(row=DATA_START_ROW, column=COLS['Expenses']).value = prev_balance

    # --- STEP H: Write Transactions (starting at Row 13) ---
    row = DATA_START_ROW + 1
    parsed_count = 0

    for raw_date, raw_desc, raw_amt in transactions:
        date_obj = parse_date(raw_date, STATEMENT_YEAR)
        amount = parse_amount(raw_amt)
        clean_name, category = get_company_and_category(raw_desc, keyword_bank)

        # Track new merchants for JSON update
        if category == 'Ask Accountant':
            cleaned = clean_description(raw_desc)
            cleaned_upper = cleaned.upper()
            if cleaned_upper and cleaned_upper not in keyword_bank:
                new_merchants.add(cleaned_upper)

        # Write Date (mm/dd/yyyy format)
        sheet.cell(row=row, column=COLS['Date']).value = date_obj
        sheet.cell(row=row, column=COLS['Date']).number_format = 'mm/dd/yyyy'

        # Write Description
        sheet.cell(row=row, column=COLS['Description']).value = clean_name

        # ROUTING LOGIC
        if category == 'Payment':
            sheet.cell(row=row, column=COLS['Pay_Back']).value = -abs(amount)
        else:
            sheet.cell(row=row, column=COLS['Expenses']).value = amount

            if category == 'Ask Accountant':
                sheet.cell(row=row, column=COLS['Ask_Accountant']).value = amount
            elif category in COLS['Categories']:
                cat_col = COLS['Categories'][category]
                sheet.cell(row=row, column=cat_col).value = amount

        row += 1
        parsed_count += 1

    # --- STEP I: Update Keywords JSON with new merchants ---
    if new_merchants:
        added = 0
        for merchant in sorted(new_merchants):
            if merchant not in keyword_bank:
                keyword_bank[merchant] = [merchant, None]
                added += 1
        if added > 0:
            save_keywords(keyword_bank)
            print(f"Added {added} new merchant(s) to {KEYWORDS_FILE}")

    wb.save(EXCEL_FILE)
    print(f"Success! Wrote Previous Balance + {parsed_count} transactions to {EXCEL_FILE}")


if __name__ == "__main__":
    parse_and_export()
